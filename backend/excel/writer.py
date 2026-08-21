"""
Professional Excel Generator (.xlsx) using openpyxl & pandas
Layout:
Row 1: Subtotal row (above Debit: sub total debit value in RED; above Credit: sub total credit value in GREEN)
Row 2: Header row (Sr No., Date, Description, Ledger, Debit, Credit, Balance) - Header colors unchanged
Row 3+: Data rows with empty Ledger column, Debit values in RED, Credit values in GREEN, numeric formatting, thin borders, and auto-filters.
Right Side Metadata (after leaving Column H blank):
- Column I (Col 9) / Column J (Col 10):
  - Row 4: Bank Name -> (empty)
  - Row 5: Account No. -> (empty)
  - Row 6: Customer Name. -> (empty)
  - Row 7: (blank)
  - Row 8: Opening Balance -> <calculated opening balance value>
  - Row 9: Closing Balance -> <calculated closing balance value>
"""
import re
from typing import List, Dict, Any, Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd
from backend.utils.logger import logger

import json
ILLEGAL_CHARACTERS_RE = re.compile(r'[\x00-\x08\x0B-\x0C\x0E-\x1F]')

def sanitize_value(val: Any) -> Any:
    if val is None:
        return ""
    if isinstance(val, (dict, list)):
        return json.dumps(val) if val else ""
    if isinstance(val, str):
        return ILLEGAL_CHARACTERS_RE.sub('', val)
    return val

def generate_excel_workbook(
    sheet_data_map: Dict[str, List[Dict[str, Any]]],
    output_filepath: str
) -> str:
    """
    Generates a professionally styled Excel workbook.
    Columns: Sr No., Date, Description, Ledger, Debit, Credit, Balance.
    Row 1: Subtotal row with Debit subtotal value (RED) and Credit subtotal value (GREEN).
    Row 2: Header row (colors untouched).
    Row 3..N+2: Transaction data rows with Debit in RED and Credit in GREEN.
    Side summary block in Column I & J (leaving Column H blank):
      - Bank Name
      - Account No.
      - Customer Name.
      - Opening Balance
      - Closing Balance
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active) # Remove default sheet

    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid") # Dark Navy Slate header
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    
    subtotal_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid") # Soft Slate Subtotal fill
    subtotal_font = Font(name="Calibri", size=11, bold=True, color="0F172A")
    subtotal_debit_font = Font(name="Calibri", size=11, bold=True, color="DC2626") # Subtotal Debit in RED
    subtotal_credit_font = Font(name="Calibri", size=11, bold=True, color="16A34A") # Subtotal Credit in GREEN
    
    regular_font = Font(name="Calibri", size=11, color="0F172A")
    label_meta_font = Font(name="Calibri", size=11, bold=False, color="000000")
    val_meta_font = Font(name="Calibri", size=11, bold=False, color="000000")
    
    debit_font = Font(name="Calibri", size=11, bold=False, color="DC2626") # Debit Value in RED
    credit_font = Font(name="Calibri", size=11, bold=False, color="16A34A") # Credit Value in GREEN
    
    alt_row_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )
    
    subtotal_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='medium', color='94A3B8')
    )

    global_sr_no = 1

    for sheet_name, transactions in sheet_data_map.items():
        safe_sheet_name = sheet_name[:30].replace("[", "").replace("]", "").replace("*", "").replace(":", "").replace("?", "").replace("/", "")
        ws = wb.create_sheet(title=safe_sheet_name or "Transactions")

        # Freeze top 2 rows (Row 1 subtotal + Row 2 header)
        ws.freeze_panes = "A3"

        # Standard requested columns: Sr No., Date, Description, Ledger, Debit, Credit, Balance
        columns_order = ["Sr No.", "Date", "Description", "Ledger", "Debit", "Credit", "Balance"]

        # Calculate numeric subtotals for Debit and Credit
        total_debit = 0.0
        total_credit = 0.0
        for tx in transactions:
            d = tx.get("Debit")
            c = tx.get("Credit")
            if d is not None:
                try:
                    d_float = float(str(d).replace(',', ''))
                    total_debit += d_float
                except (ValueError, TypeError):
                    pass
            if c is not None:
                try:
                    c_float = float(str(c).replace(',', ''))
                    total_credit += c_float
                except (ValueError, TypeError):
                    pass

        # Calculate Opening Balance & Closing Balance
        opening_balance_val: Optional[float] = None
        closing_balance_val: Optional[float] = None

        if transactions:
            # Closing balance from last transaction
            for tx in reversed(transactions):
                b = tx.get("Balance")
                if b is not None:
                    try:
                        closing_balance_val = float(str(b).replace(',', ''))
                        break
                    except (ValueError, TypeError):
                        pass

            # Opening balance from first transaction
            first_tx = transactions[0]
            first_bal = first_tx.get("Balance")
            if first_bal is not None:
                try:
                    f_bal = float(str(first_bal).replace(',', ''))
                    f_dr = 0.0
                    f_cr = 0.0
                    if first_tx.get("Debit") is not None:
                        f_dr = float(str(first_tx.get("Debit")).replace(',', ''))
                    if first_tx.get("Credit") is not None:
                        f_cr = float(str(first_tx.get("Credit")).replace(',', ''))
                    opening_balance_val = round(f_bal + f_dr - f_cr, 2)
                except (ValueError, TypeError):
                    pass

        # ----------------------------------------------------
        # ROW 1: Subtotal Row (above header)
        # ----------------------------------------------------
        row_1_values = []
        for col_name in columns_order:
            if col_name == "Debit":
                row_1_values.append(round(total_debit, 2) if total_debit > 0 else "")
            elif col_name == "Credit":
                row_1_values.append(round(total_credit, 2) if total_credit > 0 else "")
            else:
                row_1_values.append("")

        ws.append(row_1_values)

        for col_idx, col_name in enumerate(columns_order, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = subtotal_fill
            cell.border = subtotal_border
            if col_name == "Debit":
                cell.font = subtotal_debit_font
                if cell.value != "":
                    cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif col_name == "Credit":
                cell.font = subtotal_credit_font
                if cell.value != "":
                    cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.font = subtotal_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # ----------------------------------------------------
        # ROW 2: Header Row (Unchanged styling)
        # ----------------------------------------------------
        ws.append(columns_order)
        for col_idx, col_name in enumerate(columns_order, 1):
            cell = ws.cell(row=2, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        # ----------------------------------------------------
        # ROW 3+: Data Rows (Debit in RED, Credit in GREEN)
        # ----------------------------------------------------
        total_rows_needed = max(len(transactions) + 2, 9)

        for row_idx in range(3, total_rows_needed + 1):
            tx_index = row_idx - 3
            if tx_index < len(transactions):
                tx = transactions[tx_index]
                sr_num = tx.get("Sr No.") or global_sr_no
                row_data = [
                    sr_num,
                    sanitize_value(tx.get("Date", "")),
                    sanitize_value(tx.get("Description", "")),
                    sanitize_value(tx.get("Ledger", "")), # Empty Ledger column
                    tx.get("Debit"),
                    tx.get("Credit"),
                    tx.get("Balance")
                ]
                ws.append(row_data)
                global_sr_no += 1

                # Apply row styling & number formatting
                for col_idx in range(1, len(columns_order) + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.border = thin_border

                    if row_idx % 2 == 1:
                        cell.fill = alt_row_fill

                    col_name = columns_order[col_idx - 1]
                    if col_name == "Debit":
                        if cell.value is not None and isinstance(cell.value, (int, float)):
                            cell.number_format = '#,##0.00'
                            cell.alignment = Alignment(horizontal="right", vertical="center")
                            cell.font = debit_font # Red color for Debit values
                        else:
                            cell.font = regular_font
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                    elif col_name == "Credit":
                        if cell.value is not None and isinstance(cell.value, (int, float)):
                            cell.number_format = '#,##0.00'
                            cell.alignment = Alignment(horizontal="right", vertical="center")
                            cell.font = credit_font # Green color for Credit values
                        else:
                            cell.font = regular_font
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                    elif col_name == "Balance":
                        cell.font = regular_font
                        if cell.value is not None and isinstance(cell.value, (int, float)):
                            cell.number_format = '#,##0.00'
                            cell.alignment = Alignment(horizontal="right", vertical="center")
                        else:
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                    elif col_name in ["Sr No.", "Date"]:
                        cell.font = regular_font
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    elif col_name == "Ledger":
                        cell.font = regular_font
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                    else:
                        cell.font = regular_font

        # ----------------------------------------------------
        # SIDE METADATA BLOCK (Columns I & J, leaving Column H blank)
        # ----------------------------------------------------
        # Row 4: Bank Name
        # Row 5: Account No.
        # Row 6: Customer Name.
        # Row 7: (blank)
        # Row 8: Opening Balance -> opening_balance_val
        # Row 9: Closing Balance -> closing_balance_val
        side_meta_items = [
            (4, "Bank Name", ""),
            (5, "Account No.", ""),
            (6, "Customer Name.", ""),
            (8, "Opening Balance", opening_balance_val),
            (9, "Closing Balance", closing_balance_val)
        ]

        for r_idx, label, val in side_meta_items:
            cell_lbl = ws.cell(row=r_idx, column=9) # Column I
            cell_val = ws.cell(row=r_idx, column=10) # Column J

            cell_lbl.value = label
            cell_lbl.font = label_meta_font
            cell_lbl.border = thin_border
            cell_lbl.alignment = Alignment(horizontal="left", vertical="center")

            cell_val.value = val if val is not None else ""
            cell_val.font = val_meta_font
            cell_val.border = thin_border
            if isinstance(val, (int, float)):
                cell_val.number_format = '#,##0.00'
                cell_val.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell_val.alignment = Alignment(horizontal="left", vertical="center")

        # Enable Auto-filters on Header Row (Row 2, cols A to G)
        if len(transactions) > 0:
            ws.auto_filter.ref = f"A2:G{len(transactions) + 2}"

        # Auto-fit Column Widths for Columns A to G
        for col_idx in range(1, 8):
            col_letter = get_column_letter(col_idx)
            max_len = 0
            for row in range(1, len(transactions) + 3):
                cell = ws.cell(row=row, column=col_idx)
                val_str = str(cell.value or '')
                if len(val_str) > max_len:
                    max_len = len(val_str)
            ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 60)

        # Explicit widths for Ledger, Empty Column H, and Side Metadata Columns I & J
        ws.column_dimensions['D'].width = 20 # Ledger
        ws.column_dimensions['H'].width = 4  # Blank separator column after Balance
        ws.column_dimensions['I'].width = 18 # Labels (Bank Name, Opening Balance, etc.)
        ws.column_dimensions['J'].width = 18 # Values

    wb.save(output_filepath)
    logger.info(f"Excel workbook generated successfully with Side Metadata block and Ledger column: {output_filepath}")
    return output_filepath

def generate_csv(transactions: List[Dict[str, Any]], output_filepath: str) -> str:
    """Exports transactions list to CSV file with Ledger column, Subtotal row and Side metadata."""
    total_debit = 0.0
    total_credit = 0.0
    for tx in transactions:
        d = tx.get("Debit")
        c = tx.get("Credit")
        if d is not None:
            try:
                total_debit += float(str(d).replace(',', ''))
            except (ValueError, TypeError):
                pass
        if c is not None:
            try:
                total_credit += float(str(c).replace(',', ''))
            except (ValueError, TypeError):
                pass

    rows = []
    # Row 1: Subtotal values above Debit and Credit
    rows.append({
        "Sr No.": "",
        "Date": "",
        "Description": "",
        "Ledger": "",
        "Debit": round(total_debit, 2) if total_debit > 0 else "",
        "Credit": round(total_credit, 2) if total_credit > 0 else "",
        "Balance": ""
    })

    # Data Rows
    for idx, tx in enumerate(transactions, 1):
        rows.append({
            "Sr No.": tx.get("Sr No.") or idx,
            "Date": tx.get("Date", ""),
            "Description": tx.get("Description", ""),
            "Ledger": tx.get("Ledger", ""),
            "Debit": tx.get("Debit", ""),
            "Credit": tx.get("Credit", ""),
            "Balance": tx.get("Balance", "")
        })

    df = pd.DataFrame(rows)
    df.to_csv(output_filepath, index=False)
    logger.info(f"CSV file generated with Subtotal Row and Ledger column: {output_filepath}")
    return output_filepath
